from tkinter import *
from tkinter import ttk
import openpyxl as op
import math
import requests


mybook=op.load_workbook("C:\\Users\\inamz\\Desktop\\PROJECT_DATA.xlsx")

#row=sheet1.max_row
#column=sheet1.max_column
uname='1'
passwd='1'
def message2(req_marks,marks_list,subjects,reg_no,std_name,phno):
    url="https://www.fast2sms.com/dev/bulk"
    variable="Hello.\nWe are from VLITS.\nYour child "+str(std_name)+" bearing reg_no: "+reg_no+" have scored\n"
    
    for i in range(len(marks_list)):
        variable+=subjects[i]+"-"+str(marks_list[i])+"\n"
    variable+="He/She have to score\n"
    for i in range(len(marks_list)):
        variable+=subjects[i]+"-"+str(req_marks[i])+"\n"
    

    payload="sender_id=FSTSMS&message="+str(variable)+"&language=english&route=p&numbers="+str(phno)
    headers={
        "authorization":"DFQUYo4fIpXVHsiZjtSO1NgvcEPa6L2mueywzr5WRhqKlJ9dT3A2Xlx0W95g4ntYJvuOMPdrKVcpqHoG",
        "Content-Type":"application/x-www-form-urlencoded",
        "Cache-Control":"no-cache",
        }
    response=requests.request("POST",url,data=payload,headers=headers)
def message3(req_marks,marks_list,subjects,reg_no,std_name,phno):
    url="https://www.fast2sms.com/dev/bulk"
    variable="Hello.\nWe are from VLITS.\nYour child "+str(std_name)+" bearing reg_no: "+reg_no+" have scored\n"
    
    for i in range(len(subjects)):
        variable+=subjects[i]+"-"+str(marks_list[i])+"\n"
    variable+="He/She have to score \n"
    for i in range(len(subjects)):
        variable+=subjects[i]+"-"+str(req_marks[i])+"\n"
    payload="sender_id=FSTSMS&message="+str(variable)+"&language=english&route=p&numbers="+str(phno)
    headers={
        "authorization":"DFQUYo4fIpXVHsiZjtSO1NgvcEPa6L2mueywzr5WRhqKlJ9dT3A2Xlx0W95g4ntYJvuOMPdrKVcpqHoG",
        "Content-Type":"application/x-www-form-urlencoded",
        "Cache-Control":"no-cache",
        }
    response=requests.request("POST",url,data=payload,headers=headers)
def mid1(section):
    print(section)
    sheet1=mybook[section]
    row=sheet1.max_row
    column=sheet1.max_column
    for i in range(2,row+1):
        req_marks_list=[]
        marks_list=[]
        subjects=[]
        for j in range(6,11):
            if sheet1.cell(i,j).value<20:
                percent1=sheet1.cell(i,j).value * 0.8
                req_percent1=16-percent1
                required_marks1=math.ceil(req_percent1/0.2)
                percent2=sheet1.cell(i,j).value * 0.2
                req_percent2=16-percent2
                required_marks2=math.ceil(req_percent2/0.8)
                required_marks=min(required_marks1,required_marks2)
                req_marks_list.append(required_marks)
            else:
                req_marks_list.append(0)
            marks_list.append(sheet1.cell(i,j).value)
            subjects.append(sheet1.cell(1,j).value)
        message2(req_marks_list,marks_list,subjects,sheet1.cell(i,1).value,str(sheet1.cell(i,5).value),sheet1.cell(i,4).value)
def mid2(section):
    sheet1=mybook[section]
    row=sheet1.max_row
    column=sheet1.max_column
    for i in range(2,row+1):
        marks_list=[]
        req_marks=[]
        subjects=[]
        for j in range(6,11):
            
            if (sheet1.cell(i,j).value >=16 and sheet1.cell(i,j+5).value >= 16) or sheet1.cell(i,j).value>=20 or sheet1.cell(i,j+5).value>=20:
                max1=max(sheet1.cell(i,j).value,sheet1.cell(i,j+5).value)
                min1=min(sheet1.cell(i,j).value,sheet1.cell(i,j+5).value)
                total=max1*0.8 + min1*0.2
                marks_list.append(int(total))
                req_marks.append(24)
            else:
                max1=max(sheet1.cell(i,j).value,sheet1.cell(i,j+5).value)
                min1=min(sheet1.cell(i,j).value,sheet1.cell(i,j+5).value)
                total=max1*0.8 + min1*0.2
                required_marks=math.ceil(16-total)
                if required_marks<0:
                    required_marks=0
                req_marks.append(required_marks+24)
                marks_list.append(int(total))
            subjects.append(sheet1.cell(1,j).value)
        message3(req_marks,marks_list,subjects,sheet1.cell(i,1).value,sheet1.cell(i,5).value,sheet1.cell(i,4).value)
    
def message1(msg,phno):
    url="https://www.fast2sms.com/dev/bulk"
    payload="sender_id=FSTSMS&message="+str(msg)+"&language=english&route=p&numbers="+str(phno)
    headers={
        "authorization":"DFQUYo4fIpXVHsiZjtSO1NgvcEPa6L2mueywzr5WRhqKlJ9dT3A2Xlx0W95g4ntYJvuOMPdrKVcpqHoG",
        "Content-Type":"application/x-www-form-urlencoded",
        "Cache-Control":"no-cache",
        }
    response=requests.request("POST",url,data=payload,headers=headers)
def attendance(section,absenties):
    total_days=150
    completed_days=80
    sheet1=mybook[section]
    row=sheet1.max_row
    column=sheet1.max_column
    for i in range(2,row+1):
        if sheet1.cell(i,1).value[-3:] in absenties.upper():
            s="Hi, We are from VLITS\nYour child "+str(sheet1.cell(i,5).value)+" having reg_no: "+sheet1.cell(i,1).value+" is having "+str(sheet1.cell(i,2).value)+"% attendance.\n"
            if sheet1.cell(i,2).value < 75:
                for j in range(1,8):
                    a=(completed_days*(int((sheet1.cell(i,2).value))/100*7)+70*j)/(150*7)
                    if a*100>=75:
                        x=j
                        s+='\nMust attend atleast '+str(x)+' out of '+str(7)+'classes per day'
                        break
                else:
                    s+='\ndetained'        
            else:
                s+='\nAttendance Must be maintain like this'
            message1(s,sheet1.cell(i,4).value)
def absenties(event):
    def absenties_check(event):
        c,d=sections.get(),absenties.get()
        d=d.split(',')
        if c=='III_CSE_C':
            for i in d:
                if i[0]!='5':
                    Label(obj,text='re-check the details',font=('Elephant',15,'bold')).place(x=560,y=360)
                    break
            else:
                attendance(c,"".join(d))
        elif c=='III_IT':
            for i in d:
                if i[0]!='1':
                    Label(obj,text='re-check the details',font=('Elephant',15,'bold')).place(x=560,y=360)
                    break
            else:
                attendance(c,"".join(d))
        else:
            pass
    obj=Tk()
    obj.title("ABSENTIES")
    obj.geometry("1920x1080")
    #obj.configure(bg=PhotoImage(file="CIVIL_WAR-wallpaper-10895480.jpg"))
    absenties=StringVar()
    section=StringVar()
    Label(obj,text="SECTION",fg="black",font="Elephant").place(x=560,y=100)
    sections=ttk.Combobox(obj,width=25,textvariable=section)
    sections["values"]=("III_CSE_C","III_IT")
    sections.place(x=760,y=100)
    sections.current(1)
    Label(obj,text="ABSENTIES",fg="black",font="Times 15").place(x=560,y=200)
    absenties=Entry(obj,textvariable=absenties)
    absenties.place(x=760,y=200)
    b5=Button(obj,text="SUBMIT",width=8,height=1,font=16,bg="black",fg="white")
    b5.place(x=660,y=290)
    b5.bind('<Button-1>',func=absenties_check)
    obj.mainloop()
def internals(event):
    def internals_check(event):
        e,f=Exam.get(),sections.get()
        if e!='--Select--':
            if e=="Mid_1":
                mid1(f)
            elif e=="Total":
                mid2(f)
        else:
            Label(obj,text='Select Properly',font=('Helvetica',15,'bold')).place(x=600,y=80)
    obj=Tk()
    obj.title("INTERNALS")
    obj.geometry("1920x1080")
    #obj.configure(bg=PhotoImage(file="CIVIL_WAR-wallpaper-10895480.jpg"))
    Exam=StringVar()
    section=StringVar()
    Label(obj,text="SECTION",fg="black",font="Elephant").place(x=580,y=100)
    sections=ttk.Combobox(obj,width=25,textvariable=section)
    sections["values"]=("III_CSE_C","III_IT")
    sections.place(x=760,y=100)
    sections.current(0)
    Label(obj,text="Exam",fg="black",font=('Helvetica',14,'bold')).place(x=580,y=200)
    Exam=ttk.Combobox(obj,width=25,height=3,textvariable=Exam)
    Exam["values"]=('--Select--','Mid_1','Total')
    Exam.place(x=650,y=205)
    Exam.current(0)
    b6=Button(obj,text="SUBMIT",width=8,height=1,font=16,bg="black",fg="white")
    b6.place(x=660,y=290)
    b6.bind('<Button-1>',func=internals_check)
    obj.mainloop()
def login_page():
    def login_check(event):
        a,b=user_name.get(),PassWord.get()
        if a==uname and b==passwd:
            feild1=Tk()
            Label(feild1,text='login successfully',font=('Elephant',20,'bold')).pack()
            feild1.geometry('1920x1020')
            b3=Button(feild1,text="Absentees Entry",width=20,height=5,font=16,bg="black",fg="white")
            b3.place(x=400,y=250)
            b3.bind('<Button-1>',func=absenties)
            b4=Button(feild1,text="Internals",width=20,height=5,font=16,bg="black",fg="white")
            b4.place(x=800,y=250)
            b4.bind('<Button-1>',func=internals)
            feild1.mainloop()
        else:
            Label(feild,text='login unsuccessful',font=('Elephant',20,'bold')).place(x=550,y=170)
    feild=Tk()
    user_name=StringVar()
    PassWord=StringVar()
    feild.title('Login Page')
    feild.geometry('1920x1020')
    Label(text='Login Page',font=('Helvetica',40,'bold'),padx=500,pady=50,fg='black',bg='#CAF7E3').pack()
    Label(text='Username :',font=('Helvetica',13,'bold')).place(x=550,y=250)
    Entry(textvariable=user_name).place(x=660,y=250)
    Label(text='Password :',font=('Helvetica',13,'bold')).place(x=550,y=290)
    Entry(textvariable=PassWord,show='*').place(x=660,y=290)
    b2=Button(text="Submit",width=8,height=1,font=('Times 15',12,'bold'),bg="#8B8D8B",fg="white")
    b2.place(x=620,y=350)
    b2.bind('<Button-1>',func=login_check)
    feild.mainloop()
login_page()
